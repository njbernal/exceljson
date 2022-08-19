import os
from openpyxl import load_workbook
from flask import Flask, request
from flask_cors import CORS, cross_origin
from zipfile import BadZipFile

app = Flask(__name__)
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

def excel(source, ignore_blank_row = False, count = -1):
    """
    Parse a XLSX file.
    :param ignore_blank_row: Optional. 1, True or 'true' will ignore any blank rows in the file.
    :param count: Optional. The limit of how many rows to parse. If none provided, all rows will be parsed.
    :return: JSON representation of the entire file.
    """
    
    # A few validation checks
    try:
        wb = load_workbook(source)
    except BadZipFile:
        return 'Invalid file provided.'
    
    try:
        if count is None:
            count = -1
        count = int(count)
    except ValueError:
        return 'Limit parameter must be numeric.'

    if ignore_blank_row in [1, True, 'true']:
        ignore_blank_row = True

    # Let's parse this file    
    sheets = wb.sheetnames
    json_obj = []
    for sheet in sheets:
        limit = count
        worksheet = wb[sheet]
        iterable = worksheet.iter_rows()
        current = []
        while True and limit != 0:
            try:
                row = next(iterable)
                row_list = []
                for cell in row:
                    row_list.append(cell.value)
            except StopIteration:
                break
            if ignore_blank_row is True:
                check = set(row_list)
                if len(check) == 1 and None in check:
                    continue
            limit -= 1
            current.append(row_list)
        json_obj.append({sheet: current})

    return json_obj


def extension_coming(source = None, ignore = None, limit = None):
    """
    Temporary function for extensions not yet implemented that will be soon.
    """
    return 'Support for this extension coming soon.'

@app.route('/', methods=['GET', 'POST'])
@cross_origin()
def exceljson():
    if request.method == 'POST':
        ignore = request.form.get('ignore_blank_rows')
        limit = request.form.get('limit')
        source = request.files['source']
        if source.read() == b'':
            return 'No file provided. Exiting.'
        
        try:
            result = extensions[source.content_type](source, ignore, limit)
        except KeyError: 
            return f'Cannot parse a file of type {source.content_type}'
    
        # output = json.dumps(result, indent=4)
        return result

    return 'Hello there.'

# Extension dictionary to call the right parsing function
extensions = {
    "text/csv": extension_coming,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": excel
}

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8888))
    app.run(host='0.0.0.0', port=port, debug=True)
